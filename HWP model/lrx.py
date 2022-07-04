#
# -*- coding: utf-8 -*-

# Python script for running a Linny-R-model for a set of scenarios using data from Excel
# BASIC IDEA:
# (a) Scenario data are stored in one (big) Excel document (.xlsx, not .csv) that holds
#     data columns (side by side) having this structure:
#     - on row 1: column ID (short mnemonic for the variable)
#     - on row 2: name of the Linny-R variable that is to be substituted byt the column data
#     - on row 3: default value for this variable
#     - on row 4: value for this variable in time step 1 (optional)
#     - on row N: (N > 4) value for this variable in time step N - 3
# (b) Scenarios are defined in Python as a list of column IDs
# (c) For each scenario S, the Python script executes a single run of Linny-R Console (lrc.exe).
#     This will:
#     - read the columns specified in the scenario from the Excel file
#     - write this data to a CSV-bestand "S.csv" with these same columns (but w/o column ID row)
#     - execute DOS command "lrc.exe MODEL S.csv"  (MODEL being the Linny-R model name)
#     - read the columns from output file MODEL_S and convert these to an Excel workbook
#       while adding formulas for descriptive statistics at the head of each column
# The resulting set of workbooks is stored in the experiment output folder.
# The descriptive statistics from these separate files are then collected (as numbers) from 
# these files by a VisualBasic macro in a pre-configured Excel workbook viewer.xlsx.
# This same workbook will then visualize output in a compact and meaningful way.


# subprocess is needed to call the Linny-R console executable
import subprocess

# openpyxl is a Python library for accessing Excel files
from openpyxl import Workbook, load_workbook

# other libraries for supporting routines
import math, os, shutil, time

# upper case letters are used to convert numbers to Excel column labels
from string import ascii_uppercase


# path to the project directory
PROJECT_ROOT = 'D:\\M2'

# the file containing the Excel outcome viewer (copy this into the project root directory)
VIEWER_FILE = 'viewer.xlsm'

# the file containing the Linny-R model (do not add its extension .lnr !)
MODEL_FILE = 'runs2'

# assume that all scenario input data is stored in one big Excel sheet
DATA_FILE = 'shorthandwaardes14stap.xlsx'

# name of experiment (will be used to name the directory for output data files)
EXPERIMENT_NAME = 'scenarios'

# Experiment design: a list of codes V|DBC|range where:
#  V     is the code for the scenario dimension that will be shown in the viewer
#  DBC   is the column label in the Excel data file (first row)
#  range is one or more high/low or on/off symbols (typically +, o and -, but 1, 2 etc. are also OK) 
# NOTE: the dimension labels V should NOT contain range symbols (here +, o and -)
shorthand = ['efF|efF|o+', 'efE|efE|o+', 'onbP|onbPA, onbPB|+-', 'grilW| grilWPL, grilWPU, grilWQL, grilWQU, grilWRU, grilWSU|+-', 'WP|WP|o+-', 'EPB|EPB|o+-']


# auxiliary function to deal correctly with plural-s
def plural_s(number, noun):
    s = '' if int(number) == 1 else 's'
    return f'{number} {noun}{s}'


# converts seconds (as float) to a more natural time expression
def elapsed_time(sec):
    if sec < 60:
        return f'{sec:0.1f} seconds'
    if sec >= 3600:
        hrs = int(sec / 3600)
        sec -= 3600 * hrs
        hrs = plural_s(hrs, 'hour') + ', '
    else:
        hrs = ''
    mins = int(sec / 60)
    sec -= 60 * mins
    return f'{hrs}{plural_s(mins, "minute")} and {plural_s(int(sec), "second")}'


# recursive function to convert integer to Excel column label
def column_letters(n):
    if n < 26:
        return ascii_uppercase[n]
    return column_letters(math.floor(n / 26)) + ascii_uppercase[n % 26]


# recursive function that generates a full factorial scenario dict from a shorthand list
# example: ['A|Alfa, Charlie|+-', 'B|Bravo|+o-'] will result in a dictionary D of 6 scenarios
# having labels 'A+B+', 'A+Bo', 'A+B-', 'A-B+', etc. where D['A+B+'] = ['Alfa+', 'Charlie+', 'Bravo+'] etc.
def full_factorial(shorthand):
    if len(shorthand) == 0:
        return {'': []}
    ff = full_factorial(shorthand[:-1])
    sh = shorthand[-1].split('|')
    col_code = sh[0]
    col_names = [n.strip() for n in sh[1].split(',')]
    options = list(sh[2])
    ffx = {}
    for k in ff.keys():
        for o in options:
            ffx[k + col_code + o] = ff[k] + [cn + o for cn in col_names] 
    return ffx
    

# define Experiment class to do all the work
class Experiment():
    # initializa database and scenario dictionary
    def __init__(self, name, s_dict):
        self.name = name
        # use name as output directory name
        self.path = os.path.join(PROJECT_ROOT, name)
        # NOTE: protect user against overwriting an  earlier experiment!
        if os.path.exists(self.path):
            print(f'Directory {self.path} already exists.')
            print('Remove it, or change experiment name.')
            # this is a fatal error, so abort
            exit()
        try:
            # make a new directory
            os.makedirs(self.path)
        except:
            print(f'Name "{self.name}" is not usable as directory name')
            # this is a fatal error, so abort
            exit()
        # read the database
        self.rows = []
        # open the database Excel file as a workbook 
        scenario_wb = load_workbook(DATA_FILE, read_only=True, data_only=True)
        # select the active worksheet
        ws = scenario_wb['Data']
        # read all cell VALUES from this worksheet into a list of rows
        self.rows = []
        for row in ws.values:
            r = []
            for value in row:
                r.append(value)
            self.rows.append(r)
        # verify that all column IDs used in s_set exist in database
        self.scenarios = {}
        for k in s_dict.keys():
            ok = True
            for cid in s_dict[k]: 
                if not cid in self.rows[0]:
                    print(f'Unknown column ID "{cid}" in scenario {k}')
                    ok = False
            if ok:
                # if valid, add the scenario to this experiment
                self.scenarios[k] = s_dict[k]
            else:
                print(f'Ignoring scenario {k}')


    # run one scenario (specified by its ID)
    def run_one(self, scenario_id):
        if not scenario_id in self.scenarios:
            print(f'Unknown scenario "{scenario_id}"')
            return
        try:
            # select the data for the columns specified for the scenario
            csv_data = []  # list of strings to output to CSV file
            # get indices of selected columns for this scenario
            cols = [self.rows[0].index(v) for v in self.scenarios[scenario_id]]
            # first line of CSV file contains Linny-R variable names as quoted strings
            csv_data.append(';'.join([f'"{self.rows[1][c]}"' for c in cols]))
            # next lines contain numerical data
            for r in self.rows[2:]:
                csv_data.append(';'.join([f'{r[c]:0.8f}' for c in cols]))
            # output data to CSV input file for this run
            with open(f'{scenario_id}.csv', 'w') as text_file:
                text_file.write('\n'.join(csv_data))

            # execute Linny-R console
            subprocess.call(['lrc.exe', MODEL_FILE, f'{scenario_id}.csv'])

            # read data from the CSV output file
            with open(f'{MODEL_FILE}_{scenario_id}.csv', 'r') as text_file:
                csv_data = text_file.read().strip().split('\n')
            # get the number of data rows (do not count first row as it contains names)
            row_count = len(csv_data) - 1
            # create a new Excel workbook
            wb = Workbook()
            # select the first worksheet
            ws = wb.active
            # get the first line with formula names
            row = [v.strip('"') for v in csv_data[0].split(';')]
            # store it in the first row
            ws.append(row)
            # add five rows for descriptive statistics
            ranges = [f'{column_letters(i)}7:{column_letters(i)}{row_count + 6}'
                for i in range(len(row))]
            ws.append([f'=MIN({r})' for r in ranges])
            ws.append([f'=MAX({r})' for r in ranges])
            ws.append([f'=AVERAGE({r})' for r in ranges])
            ws.append([f'=STDEV({r})' for r in ranges])
            ws.append([f'=COUNTIF({r},">0")' for r in ranges])
            # also store numeric data in the Excel file
            for line in csv_data[1:]:
                # add lines as rows of floating point numbers
                row = [float(v) for v in line.split(';')]
                ws.append(row)
            # save the workbook as Excel file
            wb.save(os.path.join(self.path, f'{scenario_id}.xlsx'))
            # if successful, remove CSV files and also the LP_SOLVE and log files
            os.remove(f'{scenario_id}.csv')
            os.remove(f'{MODEL_FILE}_{scenario_id}.csv')
            os.remove(f'{MODEL_FILE}_{scenario_id}.lp')
            os.remove(f'{MODEL_FILE}_{scenario_id}.log')
        except Exception as e:
            print(f'Exception: {e}')

    # run all scenarios in this experiment
    def run_all(self):
        self.start = time.time()
        n = len(self.scenarios)
        i = 0
        for id in self.scenarios:
            i += 1 
            print(f'{time.asctime(time.localtime())} -- {id} ({i} of {n})')
            self.run_one(id)
        print(f'Experiment took {elapsed_time(time.time() - self.start)}')

"""
MAIN BODY STARTS HERE
"""
# get current directory
cwd = os.getcwd()

# change to project directory
os.chdir(PROJECT_ROOT)

# define the experiment
"""
NOTES:
(1) An experiment is a dictionary of scenarios, where each scenario is a list of 
    columnn IDs indicating that this column should be included in the Linny-R Console
    input file to overwrite that time series in the model's default data set.
(2) This Python script provides a convenient shorthand notation for creating a full factorial
    experimental design: each dimension is defined by a string 'label|column IDs|range'
    where:
    - label is a (very) short code corresponding for a Linny-R scenario variable
      NOTE: labels should NOT contain range symbols (see below) !!!
    - column IDs is a comma-separated list of the column IDs to be included;
      allowing a list is useful when several scenario variables change in the same direction
      of the dimension; typically this will be the LB and UB of products when LB = UB
    - range is a string with symbols (one character each, typically -, o and +) that are used
      as suffix for the column IDs, where o denotes the column with the "base case" values,
      and - and + the columns with low resp. high scenario values.
"""
# generate the full factorial scenario dictionary
scenario_dict = full_factorial(shorthand)

# copy the scenarios and also the labels and ranges of the dimensions
# to the Excel outcome viewer (if it exists)
print('Configuring the viewer for ', len(shorthand), ' dimensions')
if os.path.isfile(VIEWER_FILE):
    try:
        # open the viewer Excel file as a workbook 
        viewer_wb = load_workbook(VIEWER_FILE, keep_vba=True)
        # select the setup worksheet
        ws = viewer_wb['Setup']
        # write the number of dimensions
        ws['A3'] = len(shorthand)
        # write the dimensions, their range (# options), and options (1 character)
        # in their respective columns (D, E, and starting in F)
        row = 8  # start in row 8
        for dim in shorthand:
            print('Preparing for dimension ', dim)
            sd = dim.split('|')
            # ensure that dimension was properly defined
            if len(sd) == 3:
                ws[f'D{row}'] = sd[0]
                # split the range string into a list of single characters
                options = list(sd[2])
                ws[f'E{row}'] = len(options)
                # store options as string expression to allow for operators such as + and -
                col = 6
                for o in options:
                    ws.cell(row=row, column=col).value = f'="{o}"'
                    col += 1
                row += 1
            else:
                print(f'Warning: Shorthand contains invalid dimension "{dim}"')
        # also store the scenario IDs, starting again in row 8
        row = 8
        for k in scenario_dict.keys():
            ws[f'A{row}'] = k
            ws[f'B{row}'] = row - 8
            row += 1
        # and store the # of scenarios
        ws['A4'] = row - 8
        viewer_wb.save(VIEWER_FILE)
    except Exception as e:
        print('Warning: Could not configure viewer; if opened in Excel, please close it')
        print('Python error:', str(e))
else:
    print('Warning: Excel viewer not found; experiment dimensions not stored')

# create the experiment
exp = Experiment(EXPERIMENT_NAME, scenario_dict)

# run the experiment (perform a model run for each scenario)
exp.run_all()

# return to the original directory
os.chdir(cwd)

exit()
